# controllers/relatorio_controller.py
from flask import Blueprint, jsonify, send_from_directory
from services.relatorio_service import RelatorioService
from utils.export_utils import criar_pasta_exportacao
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

# Cria o Blueprint para relatórios
relatorio_bp = Blueprint('relatorio', __name__)
service = RelatorioService()

@relatorio_bp.route('/api/exportar/relatorio-diario')
def api_exportar_relatorio_diario():
    """API para exportar relatório completo do dia com totais"""
    try:
        relatorio = service.gerar_relatorio_diario()
        
        if not relatorio:
            return jsonify({
                "success": False,
                "message": "❌ Nenhuma transferência hoje para exportar"
            })
        
        # Extrair dados do relatório
        df_detalhes = relatorio['df_detalhes']
        df_totais = relatorio['df_totais']
        df_geral = relatorio['df_geral']
        data_formatada = relatorio['data_formatada']
        hora_formatada = relatorio['hora_formatada']
        
        # Criar pasta e caminho do arquivo
        pasta = criar_pasta_exportacao()
        data_exportacao = pd.Timestamp.now().strftime('%d-%m-%Y_%H-%M')
        nome_arquivo = f'relatorio_diario_{data_exportacao}.xlsx'
        caminho_arquivo = os.path.join(pasta, nome_arquivo)
        
        # Criar Excel com formatação
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            
            # Criar um DataFrame único com todas as informações
            dados_completos = []
            
            # 1. CABEÇALHO DO RELATÓRIO
            dados_completos.append(["RELATÓRIO DIÁRIO DE TRANSFERÊNCIAS"])
            dados_completos.append([f"Data: {data_formatada}"])
            dados_completos.append([f"Hora: {hora_formatada}"])
            dados_completos.append([])
            
            # 2. RESUMO GERAL
            dados_completos.append(["RESUMO"])
            dados_completos.append(["Total Transferências:", "", int(df_geral.iloc[0]['total_geral'])])
            dados_completos.append(["WhatsApp:", "", int(df_geral.iloc[0]['total_whatsapp'])])
            dados_completos.append(["Instagram:", "", int(df_geral.iloc[0]['total_instagram'])])
            dados_completos.append([])
            
            # 3. TOTAIS POR COLABORADOR
            dados_completos.append(["TOTAL"])
            dados_completos.append(["Colaborador", "Setor", "WhatsApp", "Instagram", "TOTAL"])
            
            for _, row in df_totais.iterrows():
                dados_completos.append([
                    row['colaborador'],
                    row['setor'],
                    int(row['whatsapp']),
                    int(row['instagram']),
                    int(row['total'])
                ])
            
            # Adicionar linha de totais
            dados_completos.append([
                "TOTAL GERAL",
                "",
                int(df_geral.iloc[0]['total_whatsapp']),
                int(df_geral.iloc[0]['total_instagram']),
                int(df_geral.iloc[0]['total_geral'])
            ])
            
            dados_completos.append([])
            
            # 4. DETALHES
            dados_completos.append(["DETALHES"])
            dados_completos.append(["DATA", "HORA", "COLABORADOR", "SETOR", "CANAL"])
            
            for _, row in df_detalhes.iterrows():
                dados_completos.append([
                    row['data_only'],
                    row['hora_only'],
                    row['colaborador'],
                    row['setor'],
                    row['canal']
                ])
            
            # Converter para DataFrame e exportar
            df_completo = pd.DataFrame(dados_completos)
            df_completo.to_excel(writer, sheet_name='Relatório Diário', index=False, header=False)
            
            # Formatar a planilha
            worksheet = writer.sheets['Relatório Diário']
            
            # Mesclar células
            worksheet.merge_cells('A1:E1')
            worksheet.merge_cells('A2:E2')
            worksheet.merge_cells('A3:E3')
            
            # Larguras das colunas
            col_widths = {'A': 12, 'B': 10, 'C': 20, 'D': 15, 'E': 12}
            for col, width in col_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formatar células
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=len(dados_completos), max_col=5), 1):
                for cell in row:
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=16)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif row_idx in [2, 3]:
                        cell.font = Font(bold=True, size=11)
                        cell.alignment = Alignment(horizontal='center')
                    elif row_idx in [5, 10, 12 + len(df_totais) + 1]:
                        cell.font = Font(bold=True, size=12, color="366092")
                    elif row_idx in [6, 7, 8, 11, 13 + len(df_totais) + 1]:
                        if row_idx in [11, 13 + len(df_totais) + 1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                        else:
                            cell.font = Font(bold=True)
                    elif row_idx == 11 + len(df_totais):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    elif row_idx > 13 + len(df_totais) + 1:
                        cell.alignment = Alignment(vertical='center')
                    elif row_idx > 11 and row_idx < 11 + len(df_totais) + 1:
                        cell.alignment = Alignment(vertical='center')
        
        return jsonify({
            "success": True,
            "message": f"✅ Relatório diário exportado: {nome_arquivo}",
            "arquivo": nome_arquivo,
            "quantidade": relatorio['quantidade'],
            "data": data_formatada,
            "url": f"/exportacoes/{nome_arquivo}"
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"❌ Erro na exportação do relatório: {str(e)}"
        }), 500

@relatorio_bp.route('/exportacoes/<path:filename>')
def download_arquivo(filename):
    """Rota para baixar arquivos exportados"""
    try:
        pasta = criar_pasta_exportacao()
        return send_from_directory(pasta, filename, as_attachment=True)
    except FileNotFoundError:
        return jsonify({"success": False, "message": "Arquivo não encontrado"}), 404